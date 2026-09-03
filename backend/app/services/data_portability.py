from __future__ import annotations

import csv
import io
import json
from collections import Counter
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from uuid import UUID

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.errors import AppError
from app.models.audit import AuditEvent
from app.models.control import ChangeRequest, Issue, Risk
from app.models.documents import ImportBatch, ImportStatus, ImportTarget
from app.models.finance import Expense, ExpenseStatus
from app.models.milestone import Milestone
from app.models.people import ProjectMember
from app.models.task import Task, TaskPriority, TaskStatus
from app.schemas.documents import (
    ExportDataset,
    ImportConfirmRead,
    ImportPreviewRead,
    ReportRead,
    ReportSection,
    ReportType,
)
from app.services.audit import AuditService
from app.services.authorization import AuthorizationService, Capability
from app.services.control import ControlService
from app.services.finance import FinanceService
from app.services.people import PeopleService
from app.services.projects import ProjectService
from app.services.work_planning import WorkPlanningService


class ReportingService:
    def __init__(self, session: AsyncSession, owner_user_id: UUID) -> None:
        self.session = session
        self.owner_user_id = owner_user_id
        self.audit = AuditService(session, owner_user_id)

    async def report(self, project_id: UUID, report_type: ReportType) -> ReportRead:
        authorization = AuthorizationService(self.session, self.owner_user_id)
        await authorization.require(project_id, Capability.REPORTS_GENERATE)
        can_finance = await authorization.can(project_id, Capability.FINANCE_READ)
        if report_type == ReportType.BUDGET and not can_finance:
            await authorization.require(project_id, Capability.FINANCE_READ)
        project = await ProjectService(self.session, self.owner_user_id).get(project_id)
        now = datetime.now(UTC)
        period_start = now - timedelta(days=7) if report_type == ReportType.WEEKLY else None

        work = await WorkPlanningService(self.session, self.owner_user_id).summary(project_id)
        finance = (
            await FinanceService(self.session, self.owner_user_id).analytics(project_id)
            if can_finance
            else None
        )
        control = await ControlService(self.session, self.owner_user_id).summary(project_id)
        people_service = PeopleService(self.session, self.owner_user_id)
        people = await people_service.summary(project_id)
        workloads = await people_service.workload(project_id)

        project_section = ReportSection(
            key="project",
            title="Project",
            data={
                "name": project.name,
                "code": project.code,
                "status": project.status.value,
                "priority": project.priority.value,
                "start_date": project.start_date,
                "target_end_date": project.target_end_date,
            },
        )
        delivery_section = ReportSection(
            key="delivery", title="Delivery", data=work.model_dump(mode="json")
        )
        budget_section = (
            ReportSection(key="budget", title="Budget", data=finance.model_dump(mode="json"))
            if finance is not None
            else None
        )

        risks = list(
            (
                await self.session.execute(
                    select(Risk)
                    .where(Risk.project_id == project_id)
                    .order_by((Risk.probability * Risk.impact).desc(), Risk.created_at.desc())
                    .limit(20)
                )
            ).scalars()
        )
        issues = list(
            (
                await self.session.execute(
                    select(Issue)
                    .where(Issue.project_id == project_id)
                    .order_by(Issue.created_at.desc())
                    .limit(20)
                )
            ).scalars()
        )
        changes = list(
            (
                await self.session.execute(
                    select(ChangeRequest)
                    .where(ChangeRequest.project_id == project_id)
                    .order_by(ChangeRequest.created_at.desc())
                    .limit(20)
                )
            ).scalars()
        )
        control_section = ReportSection(
            key="control",
            title="Risks and control",
            data={
                "summary": control.model_dump(mode="json"),
                "risks": [
                    {
                        "title": item.title,
                        "status": item.status.value,
                        "probability": item.probability,
                        "impact": item.impact,
                        "score": item.probability * item.impact,
                    }
                    for item in risks
                ],
                "issues": [
                    {
                        "title": item.title,
                        "status": item.status.value,
                        "priority": item.priority.value,
                    }
                    for item in issues
                ],
                "change_requests": [
                    {"title": item.title, "status": item.status.value} for item in changes
                ],
                "limits": {"records_per_type": 20},
            },
        )
        team_section = ReportSection(
            key="team",
            title="Team and workload",
            data={
                "summary": people.model_dump(mode="json"),
                "members": [item.model_dump(mode="json") for item in workloads],
            },
        )

        sections = [project_section]
        if report_type == ReportType.PROJECT_SUMMARY:
            sections.extend([delivery_section, control_section, team_section])
            if budget_section is not None:
                sections.insert(2, budget_section)
        elif report_type == ReportType.EXECUTIVE_SUMMARY:
            totals = finance.totals if finance is not None else None
            executive_data = {
                "project_status": project.status.value,
                "task_progress_percent": work.progress,
                "overdue_tasks": work.overdue_tasks,
                "upcoming_milestones": work.upcoming_milestones,
                "high_or_critical_risks": control.high_critical_risks,
                "critical_issues": control.critical_issues,
                "workload_warnings": people.workload_warning_count,
            }
            if totals is not None:
                executive_data.update(
                    {
                        "budget_utilization_percent": totals.budget_utilization,
                        "financial_status": totals.financial_status,
                    }
                )
            sections.append(
                ReportSection(
                    key="executive",
                    title="Executive summary",
                    data=executive_data,
                )
            )
        elif report_type == ReportType.BUDGET:
            assert budget_section is not None
            sections.append(budget_section)
        elif report_type == ReportType.CONTROL:
            sections.append(control_section)
        elif report_type == ReportType.TEAM:
            sections.append(team_section)
        else:
            event_rows = list(
                (
                    await self.session.execute(
                        select(AuditEvent)
                        .where(
                            AuditEvent.project_id == project_id,
                            AuditEvent.created_at >= period_start,
                            AuditEvent.created_at <= now,
                        )
                        .order_by(AuditEvent.created_at.desc())
                        .limit(200)
                    )
                ).scalars()
            )
            action_counts = Counter(item.action for item in event_rows)
            completed = int(
                (
                    await self.session.execute(
                        select(func.count(Task.id)).where(
                            Task.project_id == project_id,
                            Task.status == TaskStatus.DONE,
                            Task.updated_at >= period_start,
                            Task.updated_at <= now,
                        )
                    )
                ).scalar_one()
            )
            newly_overdue = int(
                (
                    await self.session.execute(
                        select(func.count(Task.id)).where(
                            Task.project_id == project_id,
                            Task.due_date >= period_start.date(),
                            Task.due_date < now.date(),
                            Task.status.not_in((TaskStatus.DONE, TaskStatus.CANCELLED)),
                        )
                    )
                ).scalar_one()
            )
            sections.extend(
                [
                    ReportSection(
                        key="period",
                        title="Rolling seven-day period",
                        data={
                            "period_start": period_start,
                            "period_end": now,
                            "completed_tasks": completed,
                            "newly_overdue_tasks": newly_overdue,
                            "event_count": len(event_rows),
                            "event_counts_by_action": dict(sorted(action_counts.items())),
                            "history_limit": 200,
                            "historical_data_note": (
                                "Facts reflect persisted project events available for this period."
                            ),
                        },
                    ),
                    delivery_section,
                    control_section,
                ]
            )
            if budget_section is not None:
                sections.append(budget_section)

        report = ReportRead(
            project_id=project_id,
            type=report_type,
            title=f"{project.name} — {report_type.value.replace('-', ' ').title()}",
            generated_at=now,
            period_start=period_start,
            period_end=now if period_start else None,
            sections=sections,
        )
        self.audit.record(
            project_id=project_id,
            action="report.generated",
            entity_type="project_report",
            entity_id=project_id,
            changes={"report_type": report_type.value},
        )
        await self.session.commit()
        return report

    @staticmethod
    def pdf(report: ReportRead) -> bytes:
        output = io.BytesIO()
        styles = getSampleStyleSheet()
        story = [
            Paragraph(report.title, styles["Title"]),
            Paragraph(f"Generated {report.generated_at.isoformat()}", styles["Normal"]),
            Spacer(1, 14),
        ]
        for section in report.sections:
            story.append(Paragraph(section.title, styles["Heading2"]))
            data = section.data if isinstance(section.data, dict) else {"value": section.data}
            rows = [["Field", "Value"]] + [
                [str(key).replace("_", " ").title(), str(value)] for key, value in data.items()
            ]
            table = Table(rows, colWidths=[150, 330], repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF8")),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.extend([table, Spacer(1, 12)])
        SimpleDocTemplate(output, pagesize=A4, title=report.title, author="Loris PMO").build(story)
        return output.getvalue()


class ExportService:
    MODELS = {
        ExportDataset.TASKS: (
            Task,
            [
                "id",
                "title",
                "status",
                "priority",
                "start_date",
                "due_date",
                "estimated_effort",
                "actual_effort",
                "completion_percentage",
            ],
        ),
        ExportDataset.MILESTONES: (Milestone, ["id", "title", "status", "due_date"]),
        ExportDataset.EXPENSES: (
            Expense,
            ["id", "description", "amount", "date", "supplier", "payer", "status"],
        ),
        ExportDataset.RISKS: (Risk, ["id", "title", "status", "probability", "impact"]),
        ExportDataset.ISSUES: (Issue, ["id", "title", "status", "priority"]),
        ExportDataset.CHANGES: (
            ChangeRequest,
            ["id", "title", "status", "requested_date", "estimated_delay_days", "estimated_cost"],
        ),
        ExportDataset.TEAM: (ProjectMember, ["id", "person_id", "role", "availability_percent"]),
        ExportDataset.ACTIVITY: (
            AuditEvent,
            ["id", "action", "entity_type", "entity_id", "created_at"],
        ),
    }

    def __init__(self, session: AsyncSession, owner_user_id: UUID) -> None:
        self.session = session
        self.owner_user_id = owner_user_id
        self.audit = AuditService(session, owner_user_id)

    async def rows(
        self, project_id: UUID, dataset: ExportDataset
    ) -> tuple[list[str], list[list[object]]]:
        await ProjectService(self.session, self.owner_user_id).get(project_id)
        capability = {
            ExportDataset.EXPENSES: Capability.FINANCE_READ,
            ExportDataset.ACTIVITY: Capability.AUDIT_READ,
            ExportDataset.TASKS: Capability.TASKS_READ,
            ExportDataset.MILESTONES: Capability.TASKS_READ,
            ExportDataset.RISKS: Capability.CONTROL_READ,
            ExportDataset.ISSUES: Capability.CONTROL_READ,
            ExportDataset.CHANGES: Capability.CONTROL_READ,
            ExportDataset.TEAM: Capability.PEOPLE_READ,
        }[dataset]
        await AuthorizationService(self.session, self.owner_user_id).require(project_id, capability)
        model, fields = self.MODELS[dataset]
        result = await self.session.execute(
            select(model).where(model.project_id == project_id).order_by(model.created_at)
        )
        rows = [[getattr(item, field, None) for field in fields] for item in result.scalars()]
        return fields, rows

    async def export(self, project_id: UUID, dataset: ExportDataset, fmt: str) -> bytes:
        fields, rows = await self.rows(project_id, dataset)
        if fmt == "csv":
            text = io.StringIO(newline="")
            writer = csv.writer(text)
            writer.writerow(fields)
            writer.writerows(rows)
            payload = text.getvalue().encode("utf-8-sig")
        elif fmt == "xlsx":
            book = Workbook()
            sheet = book.active
            sheet.title = dataset.value[:31]
            sheet.append(fields)
            for row in rows:
                values = []
                for value in row:
                    if hasattr(value, "value"):
                        value = value.value
                    if isinstance(value, UUID):
                        value = str(value)
                    elif isinstance(value, Decimal):
                        value = float(value)
                    elif isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False, default=str)
                    values.append(value)
                sheet.append(values)
            for column, field in enumerate(fields, start=1):
                if field in {"amount", "estimated_cost", "estimated_effort", "actual_effort"}:
                    for cell in sheet.iter_cols(
                        min_col=column, max_col=column, min_row=2
                    ).__next__():
                        cell.number_format = "0.00"
                elif field.endswith("date") or field.endswith("_at"):
                    for cell in sheet.iter_cols(
                        min_col=column, max_col=column, min_row=2
                    ).__next__():
                        cell.number_format = "yyyy-mm-dd"
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            buffer = io.BytesIO()
            book.save(buffer)
            payload = buffer.getvalue()
        else:
            raise AppError(
                code="export_format_invalid", message="Unsupported export format.", status_code=422
            )
        self.audit.record(
            project_id=project_id,
            action="export.generated",
            entity_type="project_export",
            entity_id=project_id,
            changes={"dataset": dataset.value, "format": fmt, "row_count": len(rows)},
        )
        await self.session.commit()
        return payload


class ImportService:
    MAX_ROWS = 1000

    def __init__(self, session: AsyncSession, owner_user_id: UUID) -> None:
        self.session = session
        self.owner_user_id = owner_user_id
        self.audit = AuditService(session, owner_user_id)

    @staticmethod
    def _read(data: bytes, extension: str) -> list[dict]:
        if extension == "csv":
            return list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))
        if extension == "json":
            value = json.loads(data)
            if not isinstance(value, list) or not all(isinstance(row, dict) for row in value):
                raise ValueError("JSON must contain an array of objects")
            return value
        if extension == "xlsx":
            book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheet = book.active
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                return []
            headers = [str(value or "").strip() for value in values[0]]
            return [dict(zip(headers, row, strict=False)) for row in values[1:]]
        raise ValueError("Unsupported import format")

    @staticmethod
    def _normalize(target: ImportTarget, rows: list[dict]) -> tuple[list[dict], list[dict]]:
        normalized, errors = [], []
        for number, raw in enumerate(rows[: ImportService.MAX_ROWS], start=2):
            row = {str(key).strip().lower(): value for key, value in raw.items()}
            try:
                if target == ImportTarget.TASKS:
                    title = str(row.get("title") or "").strip()
                    if not title:
                        raise ValueError("title is required")
                    item = {
                        "title": title[:240],
                        "description": str(row.get("description") or "")[:4000] or None,
                        "status": TaskStatus(str(row.get("status") or "BACKLOG").upper()).value,
                        "priority": TaskPriority(
                            str(row.get("priority") or "MEDIUM").upper()
                        ).value,
                        "start_date": ImportService._date(row.get("start_date")),
                        "due_date": ImportService._date(row.get("due_date")),
                        "estimated_effort": str(
                            ImportService._decimal(row.get("estimated_effort"), default="0")
                        ),
                        "actual_effort": str(
                            ImportService._decimal(row.get("actual_effort"), default="0")
                        ),
                        "completion_percentage": int(row.get("completion_percentage") or 0),
                        "notes": str(row.get("notes") or "")[:4000] or None,
                    }
                    if (
                        item["start_date"]
                        and item["due_date"]
                        and item["due_date"] < item["start_date"]
                    ):
                        raise ValueError("due_date precedes start_date")
                    if not 0 <= item["completion_percentage"] <= 100:
                        raise ValueError("completion_percentage must be 0-100")
                else:
                    description = str(row.get("description") or "").strip()
                    if not description:
                        raise ValueError("description is required")
                    amount = ImportService._decimal(row.get("amount"))
                    if amount <= 0:
                        raise ValueError("amount must be positive")
                    expense_date = ImportService._date(row.get("date"))
                    if expense_date is None:
                        raise ValueError("date is required")
                    item = {
                        "description": description[:300],
                        "amount": str(amount),
                        "date": expense_date,
                        "status": ExpenseStatus(str(row.get("status") or "PLANNED").upper()).value,
                        "supplier": str(row.get("supplier") or "")[:200] or None,
                        "payer": str(row.get("payer") or "")[:200] or None,
                        "notes": str(row.get("notes") or "")[:4000] or None,
                    }
                normalized.append(item)
            except (ValueError, TypeError, InvalidOperation) as exc:
                errors.append({"row": number, "message": str(exc)[:200]})
        if len(rows) > ImportService.MAX_ROWS:
            errors.append(
                {
                    "row": ImportService.MAX_ROWS + 2,
                    "message": f"Maximum {ImportService.MAX_ROWS} rows allowed",
                }
            )
        return normalized, errors

    @staticmethod
    def _date(value) -> str | None:
        if value in (None, ""):
            return None
        if isinstance(value, (datetime, date)):
            return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
        return date.fromisoformat(str(value).strip()).isoformat()

    @staticmethod
    def _decimal(value, default: str | None = None) -> Decimal:
        if value in (None, "") and default is not None:
            value = default
        return Decimal(str(value))

    async def preview(
        self, project_id: UUID, target: ImportTarget, filename: str, data: bytes
    ) -> ImportPreviewRead:
        capability = (
            Capability.FINANCE_MANAGE
            if target == ImportTarget.EXPENSES
            else Capability.TASKS_CREATE
        )
        await AuthorizationService(self.session, self.owner_user_id).require(project_id, capability)
        project = await ProjectService(self.session, self.owner_user_id).get(project_id)
        ProjectService._ensure_mutable(project)
        extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        try:
            raw = self._read(data, extension)
        except (ValueError, UnicodeError, json.JSONDecodeError) as exc:
            raise AppError(code="import_file_invalid", message=str(exc), status_code=422) from exc
        normalized, errors = self._normalize(target, raw)
        batch = ImportBatch(
            project_id=project_id,
            owner_user_id=self.owner_user_id,
            created_by_user_id=self.owner_user_id,
            target=target,
            source_format=extension,
            source_filename=filename[:255],
            normalized_rows=normalized,
            validation_errors=errors,
            status=ImportStatus.VALIDATED,
        )
        self.session.add(batch)
        await self.session.flush()
        self.audit.record(
            project_id=project_id,
            action="import.validated",
            entity_type="import_batch",
            entity_id=batch.id,
            changes={
                "target": target.value,
                "rows": len(raw),
                "valid": len(normalized),
                "errors": len(errors),
            },
        )
        await self.session.commit()
        return ImportPreviewRead(
            id=batch.id,
            project_id=project_id,
            target=target,
            source_format=extension,
            source_filename=batch.source_filename,
            row_count=len(raw),
            valid_count=len(normalized),
            errors=errors,
            preview=normalized[:20],
            can_confirm=bool(normalized) and not errors,
            status=batch.status,
        )

    async def confirm(self, project_id: UUID, batch_id: UUID) -> ImportConfirmRead:
        project = await ProjectService(self.session, self.owner_user_id).get(project_id)
        ProjectService._ensure_mutable(project)
        result = await self.session.execute(
            select(ImportBatch).where(
                ImportBatch.id == batch_id,
                ImportBatch.project_id == project_id,
                ImportBatch.owner_user_id == self.owner_user_id,
            )
        )
        batch = result.scalar_one_or_none()
        if batch is None:
            raise AppError(
                code="import_not_found", message="Import batch not found.", status_code=404
            )
        capability = (
            Capability.FINANCE_MANAGE
            if batch.target == ImportTarget.EXPENSES
            else Capability.TASKS_CREATE
        )
        await AuthorizationService(self.session, self.owner_user_id).require(project_id, capability)
        if batch.status != ImportStatus.VALIDATED:
            raise AppError(
                code="import_already_confirmed",
                message="Import was already processed.",
                status_code=409,
            )
        if batch.validation_errors or not batch.normalized_rows:
            raise AppError(
                code="import_has_errors",
                message="Correct validation errors before importing.",
                status_code=409,
            )
        for row in batch.normalized_rows:
            if batch.target == ImportTarget.TASKS:
                values = dict(row)
                values["start_date"] = (
                    date.fromisoformat(values["start_date"]) if values.get("start_date") else None
                )
                values["due_date"] = (
                    date.fromisoformat(values["due_date"]) if values.get("due_date") else None
                )
                values["estimated_effort"] = Decimal(values["estimated_effort"])
                values["actual_effort"] = Decimal(values["actual_effort"])
                self.session.add(Task(project_id=project_id, **values))
            else:
                values = dict(row)
                values["date"] = date.fromisoformat(values["date"])
                values["amount"] = Decimal(values["amount"])
                self.session.add(Expense(project_id=project_id, **values))
        batch.status = ImportStatus.COMPLETED
        self.audit.record(
            project_id=project_id,
            action="import.completed",
            entity_type="import_batch",
            entity_id=batch.id,
            changes={"target": batch.target.value, "imported_count": len(batch.normalized_rows)},
        )
        await self.session.commit()
        return ImportConfirmRead(
            batch_id=batch.id,
            target=batch.target,
            imported_count=len(batch.normalized_rows),
            status=batch.status,
        )
